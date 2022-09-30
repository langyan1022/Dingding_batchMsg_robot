# coding=GBK
# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import json
import sys

import requests
from openpyxl import load_workbook
import time
import hmac
import hashlib
import base64
import urllib.parse
import os

global webhook
global secret


class MsgLine:
    msg = ''
    mobiles = set()

    def getAtStr(self):
        str = ''
        for mobile in self.mobiles:
            str = str + "@" + mobile + " "
        return str

    def getAtArray(self):
        result = []
        for mobile in self.mobiles:
            result.append(mobile)
        return result


# ��ȡexcel����
def read_template():
    wb = load_workbook("msg-config.xlsx")
    sheet = wb.active
    global webhook
    global secret
    webhook = sheet["B1"].value
    secret = sheet["B2"].value

    maxrows = sheet.max_row
    msg_list = []
    for row in sheet.iter_rows(min_row=5, max_col=2, max_row=maxrows):
        msgLine = MsgLine()
        if len(row[0].value) == 0:
            print("����Ϣ����������")
            continue
        msgLine.msg = row[0].value
        msgLine.mobiles = set(str(row[1].value).split(','))

        msg_list.append(msgLine)

    return msg_list


def getSignStr(timestamp):
    secret_enc = secret.encode('utf-8')
    string_to_sign = '{}\n{}'.format(timestamp, secret)
    string_to_sign_enc = string_to_sign.encode('utf-8')
    hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
    sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))
    return sign


def gen_msg(msgLine):
    dic_at = {"atMobiles": msgLine.getAtArray(), "isAtAll": False}
    dic_markdown = {"title": "����", "text": '### ���� \n {} \n> {} '.format(msgLine.msg, msgLine.getAtStr())}
    dic_msg = {"at": dic_at, "msgtype": "markdown", "markdown": dic_markdown}
    # print(json.dumps(dic_msg))
    return dic_msg


def send_msg(msg_list):
    timestamp = str(round(time.time() * 1000))
    sign = getSignStr(timestamp)
    send_url = '{}&timestamp={}&sign={}'.format(webhook, timestamp, sign)
    for msgLine in msg_list:
        response = requests.post(send_url, json=gen_msg(msgLine))
        if response.status_code == 200:
            result = json.loads(response.text)
            if result['errcode'] == 0:
                print("\033[0;32;40m{} ���͸�{} �ɹ�\033[0m".format(msgLine.msg, msgLine.getAtArray()))
            else:
                print("\033[0;31;40m{} ���͸�{} ʧ�ܣ�ԭ��{}\033[0m".format(msgLine.msg, msgLine.getAtArray()),
                      result['errmsg'])
        else:
            print("\033[0;31;40m{} ���͸�{} ʧ�ܣ�ԭ��δ֪��{}\033[0m".format(msgLine.msg, msgLine.getAtArray()),
                  response.status_code)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print("��ʼ��ȡ��Ϣ����...")
    msg_list = read_template()
    if webhook == None or len(webhook) == 0 or secret == None or len(secret) == 0:
        print("\033[0;31;40m�����ļ����ô����޷�����ִ�У����������webhook���úͻ����˰�ȫsecret����\033[0m")
        os.system('pause')
        sys.exit()

    print("���ݶ�ȡ��ϣ�����һ����⵽{}����Ϣ���ݣ����������ʼ����".format(len(msg_list)))

    send_msg(msg_list)

    os.system('pause')
